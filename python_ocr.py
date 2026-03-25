"""img/ 폴더 이미지 → 흰 박스 크롭 → OCR(Upstage) → LLM(Gemini) 구조화 → 엑셀 + 작업로그."""

import base64
import io
import json
import os
import re
import ssl
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage
from PIL.ExifTags import TAGS, GPSTAGS

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env", override=True)

def _env(key: str) -> str:
    v = os.environ.get(key, "").strip()
    if not v:
        raise RuntimeError(f"{key}가 설정되지 않았습니다. .env 또는 Streamlit Secrets에 추가하세요.")
    return v


UPSTAGE_API_KEY = os.environ.get("UPSTAGE_API_KEY", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
IMG_DIR = ROOT / "img"
RESULT_DIR = IMG_DIR / "result"

ssl._create_default_https_context = ssl._create_unverified_context

# ---------------------------------------------------------------------------
# 0) 흰 박스(보드판 표) 감지 → 크롭
# ---------------------------------------------------------------------------
MIN_AREA_RATIO = 0.03  # 전체 이미지 대비 최소 면적 비율
MAX_AREA_RATIO = 0.85  # 너무 크면 배경 전체


def crop_white_box(image_bytes: bytes) -> bytes:
    """흰색 직사각형 영역을 찾아 크롭한 JPEG bytes 반환. 못 찾으면 원본 그대로."""
    arr = np.frombuffer(image_bytes, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        return image_bytes

    h, w = img.shape[:2]
    total_area = h * w

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 흰색 영역 추출 (밝은 픽셀 → 이진화)
    _, thresh = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY)

    # 노이즈 제거 후 닫힘 연산으로 흰 영역 합치기
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 15))
    closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=3)

    contours, _ = cv2.findContours(closed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    best = None
    best_area = 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        ratio = area / total_area
        if ratio < MIN_AREA_RATIO or ratio > MAX_AREA_RATIO:
            continue
        peri = cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, 0.02 * peri, True)
        # 사각형(4~6꼭짓점)이거나 boundingRect 근사
        x, y, bw, bh = cv2.boundingRect(cnt)
        rect_area = bw * bh
        fill_ratio = area / rect_area if rect_area > 0 else 0
        # 직사각형에 가까운지 (면적 채움률 70% 이상)
        if fill_ratio > 0.7 and area > best_area:
            best = (x, y, bw, bh)
            best_area = area

    if best is None:
        print("   [크롭] 흰 박스를 찾지 못함 → 원본 사용")
        return image_bytes

    x, y, bw, bh = best
    # 약간의 여백 추가
    pad = 5
    x1 = max(0, x - pad)
    y1 = max(0, y - pad)
    x2 = min(w, x + bw + pad)
    y2 = min(h, y + bh + pad)
    cropped = img[y1:y2, x1:x2]
    print(f"   [크롭] 흰 박스 감지: ({x1},{y1})→({x2},{y2}), "
          f"크기 {x2-x1}x{y2-y1} (원본 {w}x{h})")

    ok, buf = cv2.imencode(".jpg", cropped, [cv2.IMWRITE_JPEG_QUALITY, 95])
    return buf.tobytes() if ok else image_bytes


# ---------------------------------------------------------------------------
# 1) Upstage Document Parse — OCR 텍스트 추출
# ---------------------------------------------------------------------------
import requests as _req
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

OCR_URL = "https://api.upstage.ai/v1/document-digitization"


def ocr_extract(image_bytes: bytes, filename: str) -> tuple[str, int]:
    """OCR 텍스트와 사용 페이지 수를 반환."""
    api_key = UPSTAGE_API_KEY or _env("UPSTAGE_API_KEY")
    resp = _req.post(
        OCR_URL,
        headers={"Authorization": f"Bearer {api_key}"},
        files={"document": (filename, image_bytes, "image/jpeg")},
        data={
            "model": "document-parse",
            "ocr": "force",
            "output_formats": "['text']",
            "coordinates": "false",
        },
        timeout=120,
        verify=False,
    )
    resp.raise_for_status()
    data = resp.json()
    content = data.get("content", {})
    pages = data.get("usage", {}).get("pages", 1)
    text = content.get("text") or content.get("markdown") or ""
    return text, pages


# ---------------------------------------------------------------------------
# 2) LLM (Gemini) — OCR 텍스트를 건설 보드 JSON으로 구조화
# ---------------------------------------------------------------------------
PROMPT = """너는 건설 현장 데이터 교정 AI야. OCR로 추출된 텍스트 중 인식 오류(예: 숫자 오기입, 필체에 따른 글자 깨짐)를 문맥에 맞게 수정하고, 특히 건설 전문 용어 사전을 기반으로 오타를 자동으로 교정해.

다음은 건설 현장 동산보드판에서 OCR로 추출한 원문이다.
아래 키만 사용하여 JSON 객체 **하나**만 출력해라. 값이 없으면 빈 문자열 ""을 넣어라.
JSON 바깥에 설명·마크다운·코드펜스를 쓰지 마라.

키:
- project_name: 공사명
- category: 공종
- location: 위치
- details: 내용

OCR 원문:
---
{ocr_text}
---
"""

REQUIRED_KEYS = ("project_name", "category", "location", "details")

GEMINI_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.5-flash:generateContent"
)


def llm_structure(ocr_text: str, max_retries: int = 5) -> dict:
    import time
    body = {
        "contents": [{"parts": [{"text": PROMPT.format(ocr_text=ocr_text or "(빈 텍스트)")}]}],
        "generationConfig": {
            "temperature": 0.1,
            "responseMimeType": "application/json",
        },
    }
    api_key = GEMINI_API_KEY or _env("GEMINI_API_KEY")
    for attempt in range(max_retries):
        resp = _req.post(
            GEMINI_URL,
            params={"key": api_key},
            json=body,
            timeout=120,
            verify=False,
        )
        if resp.status_code == 429:
            wait = 15 * (attempt + 1)
            print(f"   [429] 요청 제한 → {wait}초 대기 후 재시도 ({attempt+1}/{max_retries})")
            time.sleep(wait)
            continue
        resp.raise_for_status()
        break
    else:
        resp.raise_for_status()
    resp_json = resp.json()
    candidate = resp_json["candidates"][0]
    raw = candidate["content"]["parts"][0]["text"].strip()

    usage = resp_json.get("usageMetadata", {})
    input_tokens = usage.get("promptTokenCount", 0)
    output_tokens = usage.get("candidatesTokenCount", 0)

    m = re.match(r"^```(?:json)?\s*\r?\n(.*?)\r?\n```\s*$", raw, re.DOTALL | re.IGNORECASE)
    if m:
        raw = m.group(1).strip()
    data = json.loads(raw)
    result = {k: str(data.get(k, "")) for k in REQUIRED_KEYS}
    result["_input_tokens"] = input_tokens
    result["_output_tokens"] = output_tokens
    return result


# ---------------------------------------------------------------------------
# 3-0) 사진 EXIF 메타데이터 추출
# ---------------------------------------------------------------------------
def _dms_to_decimal(dms, ref: str) -> float:
    """GPS 도분초(DMS) → 십진 좌표 변환."""
    d, m, s = [float(v) for v in dms]
    dec = d + m / 60 + s / 3600
    if ref in ("S", "W"):
        dec = -dec
    return dec


def extract_exif_meta(image_bytes: bytes) -> dict:
    """EXIF에서 촬영일시·GPS 좌표를 추출. 없으면 빈 문자열."""
    result = {"photo_date": "", "photo_location": ""}
    try:
        with PILImage.open(io.BytesIO(image_bytes)) as im:
            exif = im._getexif()
            if not exif:
                return result

            for tag_id, val in exif.items():
                tag = TAGS.get(tag_id, "")
                if tag in ("DateTimeOriginal", "DateTimeDigitized", "DateTime"):
                    if not result["photo_date"]:
                        result["photo_date"] = str(val).replace(":", "-", 2)

                if tag == "GPSInfo":
                    gps = {}
                    for k, v in val.items():
                        gps[GPSTAGS.get(k, k)] = v
                    try:
                        lat = _dms_to_decimal(
                            gps["GPSLatitude"], gps.get("GPSLatitudeRef", "N")
                        )
                        lon = _dms_to_decimal(
                            gps["GPSLongitude"], gps.get("GPSLongitudeRef", "E")
                        )
                        result["photo_location"] = f"{lat:.6f}, {lon:.6f}"
                    except (KeyError, TypeError, ValueError):
                        pass
    except Exception:
        pass
    return result


# ---------------------------------------------------------------------------
# 3) 엑셀 저장 — 월별 누적 (JSON + 썸네일 파일 → 엑셀 재생성)
# ---------------------------------------------------------------------------
HEADERS = ["사진", "파일명", "수행일시", "촬영일시", "촬영위치", "공사명", "공종", "위치", "내용"]
DATA_KEYS = ["project_name", "category", "location", "details"]
META_KEYS = ["added_at", "photo_date", "photo_location"]
THUMB_MAX = (120, 120)
ROW_HEIGHT = 95
THUMBS_DIR = RESULT_DIR / "thumbs"


def make_thumbnail(image_bytes: bytes) -> bytes:
    with PILImage.open(io.BytesIO(image_bytes)) as im:
        rgb = im.convert("RGB")
        rgb.thumbnail(THUMB_MAX, PILImage.Resampling.LANCZOS)
        buf = io.BytesIO()
        rgb.save(buf, format="PNG")
        return buf.getvalue()


def _monthly_paths() -> tuple[Path, Path]:
    """현재 월의 엑셀·JSON 경로 반환."""
    month = datetime.now().strftime("%Y-%m")
    return RESULT_DIR / f"{month}.xlsx", RESULT_DIR / f"{month}.json"


def append_monthly(new_rows: list[dict]) -> Path:
    """새 행을 월별 데이터에 누적하고, 엑셀을 전체 재생성하여 반환."""
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    THUMBS_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_path, json_path = _monthly_paths()

    existing: list[dict] = []
    if json_path.exists():
        existing = json.loads(json_path.read_text(encoding="utf-8"))

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    for idx, row in enumerate(new_rows):
        thumb_name = f"{ts}_{idx}_{row['filename']}.png"
        if row.get("image_bytes"):
            thumb_bytes = make_thumbnail(row["image_bytes"])
            (THUMBS_DIR / thumb_name).write_bytes(thumb_bytes)

        meta = extract_exif_meta(row.get("image_bytes", b""))
        existing.append({
            "filename": row["filename"],
            "structured": row.get("structured", {}),
            "thumb_file": thumb_name,
            "added_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "photo_date": meta.get("photo_date", ""),
            "photo_location": meta.get("photo_location", ""),
        })

    json_path.write_text(
        json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    try:
        _rebuild_excel(existing, xlsx_path)
    except PermissionError:
        alt = xlsx_path.with_name(xlsx_path.stem + "_tmp.xlsx")
        print(f"   [경고] {xlsx_path.name} 잠김 → {alt.name} 으로 저장")
        _rebuild_excel(existing, alt)
        xlsx_path = alt
    return xlsx_path


def _rebuild_excel(entries: list[dict], output_path: Path) -> None:
    """JSON 데이터 + 저장된 썸네일로 엑셀을 처음부터 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = "동산보드 OCR 결과"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.column_dimensions["A"].width = 18   # 사진
    ws.column_dimensions["B"].width = 14   # 파일명
    ws.column_dimensions["C"].width = 20   # 수행일시
    ws.column_dimensions["D"].width = 20   # 촬영일시
    ws.column_dimensions["E"].width = 24   # 촬영위치
    for col_letter in ("F", "G", "H", "I"):
        ws.column_dimensions[col_letter].width = 28

    for ri, entry in enumerate(entries, 2):
        thumb_path = THUMBS_DIR / entry.get("thumb_file", "")
        if thumb_path.is_file():
            try:
                img = XlImage(str(thumb_path))
                ws.add_image(img, f"A{ri}")
            except Exception:
                ws.cell(row=ri, column=1, value="(이미지 실패)")
        else:
            ws.cell(row=ri, column=1, value="(없음)")

        ws.cell(row=ri, column=2, value=entry.get("filename", ""))
        ws.cell(row=ri, column=3, value=entry.get("added_at", ""))
        ws.cell(row=ri, column=4, value=entry.get("photo_date", ""))
        ws.cell(row=ri, column=5, value=entry.get("photo_location", ""))
        structured = entry.get("structured", {})
        for ci, k in enumerate(DATA_KEYS, 6):
            ws.cell(row=ri, column=ci, value=structured.get(k, ""))

        ws.row_dimensions[ri].height = ROW_HEIGHT
        for ci in range(1, len(HEADERS) + 1):
            ws.cell(row=ri, column=ci).alignment = Alignment(
                vertical="center", wrap_text=True
            )

    wb.save(output_path)


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# 4) 작업 로그 — work_log.txt 누적 기록
# ---------------------------------------------------------------------------
UPSTAGE_OCR_PRICE_PER_PAGE = 0.01        # USD
GEMINI_INPUT_PRICE_PER_1M  = 0.15        # USD (gemini-2.5-flash input)
GEMINI_OUTPUT_PRICE_PER_1M = 0.60        # USD (gemini-2.5-flash output)


def write_work_log(
    num_images: int,
    ocr_pages: int,
    gemini_input_tokens: int,
    gemini_output_tokens: int,
    output_file: str,
) -> None:
    ocr_cost = ocr_pages * UPSTAGE_OCR_PRICE_PER_PAGE
    gemini_input_cost = gemini_input_tokens / 1_000_000 * GEMINI_INPUT_PRICE_PER_1M
    gemini_output_cost = gemini_output_tokens / 1_000_000 * GEMINI_OUTPUT_PRICE_PER_1M
    gemini_cost = gemini_input_cost + gemini_output_cost
    total_cost = ocr_cost + gemini_cost

    log_path = RESULT_DIR / "work_log.txt"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ocr_model = "document-parse"
    llm_model = GEMINI_URL.split("/models/")[1].split(":")[0]

    entry = (
        f"{'=' * 60}\n"
        f"실행 일시       : {now}\n"
        f"처리 사진 수    : {num_images}장\n"
        f"출력 파일       : {output_file}\n"
        f"─── OCR (Upstage) ───\n"
        f"  모델          : {ocr_model}\n"
        f"  페이지 수     : {ocr_pages}\n"
        f"  비용          : ${ocr_cost:.4f}\n"
        f"─── LLM (Gemini) ───\n"
        f"  모델          : {llm_model}\n"
        f"  입력 토큰     : {gemini_input_tokens:,}\n"
        f"  출력 토큰     : {gemini_output_tokens:,}\n"
        f"  합계 토큰     : {gemini_input_tokens + gemini_output_tokens:,}\n"
        f"  비용          : ${gemini_cost:.6f}\n"
        f"─── 합계 ───\n"
        f"  예상 총 비용  : ${total_cost:.4f}\n"
        f"{'=' * 60}\n\n"
    )

    with open(log_path, "a", encoding="utf-8") as f:
        f.write(entry)
    print(f"[로그] work_log.txt 에 기록 완료")


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    RESULT_DIR.mkdir(parents=True, exist_ok=True)

    images = sorted(IMG_DIR.glob("*.*"))
    images = [f for f in images if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".bmp", ".webp", ".gif", ".tiff", ".tif")]
    print(f"이미지 {len(images)}장 발견: {[f.name for f in images]}\n")

    rows = []
    total_ocr_pages = 0
    total_gemini_input = 0
    total_gemini_output = 0

    for idx, img_path in enumerate(images, 1):
        print(f"[{idx}/{len(images)}] {img_path.name}")
        img_bytes = img_path.read_bytes()

        print("   흰 박스 크롭 중...")
        cropped = crop_white_box(img_bytes)

        print("   OCR 추출 중...")
        ocr_text, pages = ocr_extract(cropped, img_path.name)
        total_ocr_pages += pages
        safe = ocr_text.encode("utf-8", errors="replace").decode("utf-8")
        print(f"   OCR 결과 ({len(ocr_text)}자, {pages}페이지): {safe[:120]}...")

        print("   LLM 구조화 중...")
        structured = llm_structure(ocr_text)
        total_gemini_input += structured.pop("_input_tokens", 0)
        total_gemini_output += structured.pop("_output_tokens", 0)
        for k in REQUIRED_KEYS:
            label = dict(zip(REQUIRED_KEYS, ("공사명", "공종", "위치", "내용")))[k]
            v = str(structured.get(k, "")).encode("utf-8", errors="replace").decode("utf-8")
            print(f"     {label}: {v}")

        rows.append({
            "filename": img_path.name,
            "image_bytes": img_bytes,
            "structured": structured,
        })
        print()

    print("엑셀 월별 누적 저장 중...")
    out_path = append_monthly(rows)
    print(f"[완료] {out_path.name} 저장됨 ({out_path.stat().st_size:,} bytes)")

    write_work_log(
        num_images=len(images),
        ocr_pages=total_ocr_pages,
        gemini_input_tokens=total_gemini_input,
        gemini_output_tokens=total_gemini_output,
        output_file=out_path.name,
    )
